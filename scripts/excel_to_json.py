#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel → JSON 変換スクリプト
SenYouAI Studio用

使い方:
  python excel_to_json.py data_template.xlsx

必要なライブラリ:
  pip install openpyxl
"""

import json
import sys
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl がインストールされていません")
    print("以下のコマンドでインストールしてください:")
    print("  pip install openpyxl")
    sys.exit(1)


def read_excel_to_json(excel_path):
    """Excelファイルを読み込んでJSONファイルを生成"""
    
    wb = load_workbook(excel_path)
    output_dir = Path("data")
    output_dir.mkdir(exist_ok=True)
    
    print(f"📊 Excelファイル読み込み: {excel_path}")
    print("-" * 60)
    
    # ========== Site情報 ==========
    if "Site" in wb.sheetnames:
        ws = wb["Site"]
        site_data = {
            "title": ws["B2"].value or "SenYouAI Studio / 愛玩王姫 Official",
            "tagline": ws["B3"].value or "AIとあなたで育てるバーチャルプロジェクト",
            "theme": ws["B4"].value or "dark",
            "season": ws["B5"].value or "default",
            "nav": [
                {"id": "home", "label": "Home"},
                {"id": "music", "label": "Music"},
                {"id": "novels", "label": "Novel"},
                {"id": "stamps", "label": "LINE"},
                {"id": "about", "label": "About"}
            ]
        }
        with open(output_dir / "site.json", "w", encoding="utf-8") as f:
            json.dump(site_data, f, ensure_ascii=False, indent=2)
        print("✅ site.json 生成完了")
    
    # ========== Artists情報 ==========
    if "Artists" in wb.sheetnames:
        ws = wb["Artists"]
        artists_items = []
        
        for row in range(3, ws.max_row + 1):  # 3行目からデータ開始
            artist_id = ws[f"A{row}"].value
            if not artist_id:
                break
                
            artist = {
                "id": artist_id,
                "name": ws[f"B{row}"].value or "",
                "role": ws[f"C{row}"].value or "",
                "cover": ws[f"D{row}"].value or "",
                "bio": ws[f"E{row}"].value or ""
            }
            
            # アーティストページURL（I列）
            if ws[f"I{row}"].value:
                artist["artistPageUrl"] = ws[f"I{row}"].value
            
            # SpotifyアーティストURL（J列）
            if ws[f"J{row}"].value:
                artist["spotifyArtistUrl"] = ws[f"J{row}"].value
            
            # プレイリストリンク（オプション）
            playlists = {}
            if ws[f"F{row}"].value:
                playlists["spotify"] = ws[f"F{row}"].value
            if ws[f"G{row}"].value:
                playlists["youtubeMusic"] = ws[f"G{row}"].value
            if ws[f"H{row}"].value:
                playlists["amazonMusic"] = ws[f"H{row}"].value
            if playlists:
                artist["playlists"] = playlists
            
            artists_items.append(artist)
        
        with open(output_dir / "artists.json", "w", encoding="utf-8") as f:
            json.dump({"items": artists_items}, f, ensure_ascii=False, indent=2)
        print(f"✅ artists.json 生成完了 ({len(artists_items)}件)")
    
    # ========== Music情報 ==========
    if "Music" in wb.sheetnames:
        ws = wb["Music"]
        music_items = []
        
        for row in range(3, ws.max_row + 1):  # 3行目からデータ開始
            song_id = ws[f"A{row}"].value
            if not song_id:
                break
                
            song = {
                "id": song_id,
                "title": ws[f"B{row}"].value or "",
                "artistId": ws[f"C{row}"].value or "",
                "releaseDate": ws[f"D{row}"].value or "",
                "status": ws[f"E{row}"].value or "released",
                "cover": ws[f"F{row}"].value or "",
                "lyricsPreview": ws[f"G{row}"].value or "",
                "lyrics": ws[f"H{row}"].value or "",
                "note": ws[f"I{row}"].value or ""
            }
            
            # タグ（カンマ区切り）
            tags_str = ws[f"J{row}"].value
            if tags_str:
                song["tags"] = [t.strip() for t in str(tags_str).split(",")]
            else:
                song["tags"] = []
            
            # リンク
            links = {}
            if ws[f"K{row}"].value:
                links["YouTube"] = ws[f"K{row}"].value
            if ws[f"L{row}"].value:
                links["Spotify"] = ws[f"L{row}"].value
            if ws[f"M{row}"].value:
                links["Apple Music"] = ws[f"M{row}"].value
            song["links"] = links
            
            # Spotify埋め込みURL（N列）
            if ws[f"N{row}"].value:
                song["spotifyEmbed"] = ws[f"N{row}"].value
            else:
                song["spotifyEmbed"] = ""
            
            music_items.append(song)
        
        music_data = {
            "sections": [
                {
                    "title": "Singles",
                    "items": music_items
                }
            ]
        }
        with open(output_dir / "music.json", "w", encoding="utf-8") as f:
            json.dump(music_data, f, ensure_ascii=False, indent=2)
        print(f"✅ music.json 生成完了 ({len(music_items)}件)")
    
    # ========== Novels情報 ==========
    if "Novels" in wb.sheetnames:
        ws = wb["Novels"]
        novel_items = []
        
        for row in range(3, ws.max_row + 1):
            novel_id = ws[f"A{row}"].value
            if not novel_id:
                break
                
            novel = {
                "id": novel_id,
                "title": ws[f"B{row}"].value or "",
                "subtitle": ws[f"C{row}"].value or "",
                "description": ws[f"D{row}"].value or ""
            }
            
            # リンク
            links = {}
            if ws[f"E{row}"].value:
                links["narou"] = ws[f"E{row}"].value
            if ws[f"F{row}"].value:
                links["kindle"] = ws[f"F{row}"].value
            if ws[f"G{row}"].value:
                links["other"] = ws[f"G{row}"].value
            novel["links"] = links
            
            novel_items.append(novel)
        
        with open(output_dir / "novels.json", "w", encoding="utf-8") as f:
            json.dump({"items": novel_items}, f, ensure_ascii=False, indent=2)
        print(f"✅ novels.json 生成完了 ({len(novel_items)}件)")
    
    # ========== News情報 ==========
    if "News" in wb.sheetnames:
        ws = wb["News"]
        news_items = []
        
        for row in range(3, ws.max_row + 1):  # 3行目からデータ開始
            date = ws[f"A{row}"].value
            if not date:
                break
                
            news = {
                "date": str(date) if date else "",
                "title": ws[f"B{row}"].value or "",
                "description": ws[f"C{row}"].value or "",
                "link": ws[f"D{row}"].value or "",
                "icon": ws[f"E{row}"].value or "📢"
            }
            
            news_items.append(news)
        
        with open(output_dir / "news.json", "w", encoding="utf-8") as f:
            json.dump({"items": news_items}, f, ensure_ascii=False, indent=2)
        print(f"✅ news.json 生成完了 ({len(news_items)}件)")
    
    # ========== Stamps情報 ==========
    if "Stamps" in wb.sheetnames:
        ws = wb["Stamps"]
        stamp_items = []
        
        for row in range(3, ws.max_row + 1):
            stamp_id = ws[f"A{row}"].value
            if not stamp_id:
                break
                
            stamp = {
                "id": stamp_id,
                "title": ws[f"B{row}"].value or "",
                "description": ws[f"C{row}"].value or "",
                "cover": ws[f"D{row}"].value or "",
                "listUrl": ws[f"E{row}"].value or "",
                "detailUrl": ws[f"F{row}"].value or ""
            }
            
            # タグ（カンマ区切り）
            tags_str = ws[f"G{row}"].value
            if tags_str:
                stamp["tags"] = [t.strip() for t in str(tags_str).split(",")]
            else:
                stamp["tags"] = []
            
            stamp_items.append(stamp)
        
        with open(output_dir / "stamps.json", "w", encoding="utf-8") as f:
            json.dump({"items": stamp_items}, f, ensure_ascii=False, indent=2)
        print(f"✅ stamps.json 生成完了 ({len(stamp_items)}件)")
    
    print("-" * 60)
    print(f"🎉 全JSONファイル生成完了！ (data/フォルダに保存)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使い方: python excel_to_json.py <Excelファイル名>")
        print("例: python excel_to_json.py data_template.xlsx")
        sys.exit(1)
    
    excel_file = sys.argv[1]
    if not Path(excel_file).exists():
        print(f"ERROR: ファイルが見つかりません: {excel_file}")
        sys.exit(1)
    
    read_excel_to_json(excel_file)
